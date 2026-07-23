"""
run_appium_e2e.py — Appium Mobile E2E Test Runner
=======================================================
VoiceMail AI Android App — Complete End-to-End Test Suite

Usage:
  # Full Android Appium run (requires Appium server + connected device/emulator)
  python tests/appium/run_appium_e2e.py

  # Desktop Chrome fallback (no Appium needed — for local dev / CI fallback)
  BASE_URL=http://localhost:3000 python tests/appium/run_appium_e2e.py

  # Custom Appium server URL
  APPIUM_URL=http://localhost:4723 python tests/appium/run_appium_e2e.py

Reports saved to:
  Appium Test Results/
  ├── Excel/Appium_Mobile_Test_Report.xlsx
  ├── HTML/mobile_execution_report.html
  ├── Screenshots/*.png
  ├── Logs/appium_e2e.log
  └── Summary/mobile_summary.md
"""

import os
import sys
import time
import json
import logging
import requests
from datetime import datetime

# Add tests/appium to path for POM imports
APPIUM_ROOT = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, APPIUM_ROOT)

# ── Configuration ─────────────────────────────────────────────────────────────
BASE_URL    = os.environ.get("BASE_URL", "http://localhost:3000").rstrip("/")
APPIUM_URL  = os.environ.get("APPIUM_URL", "http://localhost:4723")
DEVICE_NAME = os.environ.get("DEVICE_NAME", "Android Emulator")
PLATFORM    = os.environ.get("PLATFORM_VERSION", "13.0")
GITHUB_RUN  = os.environ.get("GITHUB_RUN_NUMBER", "Local")

# ── Output Directories ────────────────────────────────────────────────────────
RESULTS_ROOT  = "Appium Test Results"
EXCEL_DIR     = os.path.join(RESULTS_ROOT, "Excel")
HTML_DIR      = os.path.join(RESULTS_ROOT, "HTML")
SCREENSHOT_DIR= os.path.join(RESULTS_ROOT, "Screenshots")
LOGS_DIR      = os.path.join(RESULTS_ROOT, "Logs")
SUMMARY_DIR   = os.path.join(RESULTS_ROOT, "Summary")

for d in [EXCEL_DIR, HTML_DIR, SCREENSHOT_DIR, LOGS_DIR, SUMMARY_DIR]:
    os.makedirs(d, exist_ok=True)

# ── Logging Setup ─────────────────────────────────────────────────────────────
LOG_FILE = os.path.join(LOGS_DIR, "appium_e2e.log")
import io
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace"))
    ]
)
log = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# Driver Setup
# ─────────────────────────────────────────────────────────────────────────────
def create_appium_driver():
    """
    Attempt to create an Appium driver (Android Chrome).
    Falls back to desktop Selenium Chrome with a mobile viewport on failure.
    """
    try:
        import socket
        from urllib.parse import urlparse
        parsed = urlparse(APPIUM_URL)
        host = parsed.hostname or "localhost"
        port = parsed.port or 4723

        # Fast socket ping check before calling urllib3
        s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        s.settimeout(1.0)
        res = s.connect_ex((host, port))
        s.close()

        if res != 0:
            log.info(f"[INFO] Appium server port {port} not listening. Using Chrome mobile emulation driver.")
            return create_fallback_driver(), "Chrome (Mobile Emulation Mode)"

        from appium import webdriver as appium_webdriver
        try:
            from appium.options.android.uiautomator2.base import UiAutomator2Options
        except ImportError:
            try:
                from appium.options import AppiumOptions as UiAutomator2Options
            except ImportError:
                from selenium.webdriver.chrome.options import Options as UiAutomator2Options

        options = UiAutomator2Options()
        try:
            options.platform_name         = "Android"
            options.platform_version      = PLATFORM
            options.device_name           = DEVICE_NAME
            options.browser_name          = "Chrome"
            options.no_reset              = True
            options.full_reset            = False
            options.automation_name       = "UIAutomator2"
        except AttributeError:
            options = {"platformName": "Android", "platformVersion": PLATFORM,
                       "deviceName": DEVICE_NAME, "browserName": "Chrome",
                       "noReset": True, "automationName": "UIAutomator2"}

        log.info(f"Connecting to Appium at {APPIUM_URL} ...")
        driver = appium_webdriver.Remote(APPIUM_URL, options=options)
        driver.implicitly_wait(5)
        log.info(f"[OK] Appium driver connected - Device: {DEVICE_NAME}, Platform: Android {PLATFORM}")
        return driver, "Appium (Android Chrome)"

    except Exception as e:
        log.warning(f"[WARN] Appium driver fallback ({type(e).__name__}: {e}). Using Chrome mobile emulation mode.")
        return create_fallback_driver(), "Chrome (Mobile Emulation Mode)"

def create_fallback_driver():
    """Selenium Chrome with mobile device emulation (360×800 — standard Android)."""
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.chrome.service import Service

    opts = Options()
    opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--window-size=360,800")
    opts.add_experimental_option("mobileEmulation", {
        "deviceMetrics": {"width": 360, "height": 800, "pixelRatio": 3.0},
        "userAgent": (
            "Mozilla/5.0 (Linux; Android 13; Pixel 7) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/120.0.0.0 Mobile Safari/537.36"
        )
    })

    driver = webdriver.Chrome(options=opts)
    driver.implicitly_wait(10)
    return driver


# ─────────────────────────────────────────────────────────────────────────────
# Connectivity Check
# ─────────────────────────────────────────────────────────────────────────────
def wait_for_server(url: str, retries: int = 12, delay: int = 5):
    ping = f"{url}/api/auth/ping"
    log.info(f"Waiting for server at {ping} ...")
    for i in range(retries):
        try:
            r = requests.get(ping, timeout=5)
            if r.status_code == 200:
                log.info("[OK] Server is ready!")
                return True
        except Exception:
            pass
        log.info(f"  Attempt {i+1}/{retries} - waiting {delay}s ...")
        time.sleep(delay)
    log.error(f"[FAIL] Server not available after {retries} retries.")
    return False


# ─────────────────────────────────────────────────────────────────────────────
# Test Execution
# ─────────────────────────────────────────────────────────────────────────────
def run_all_tests(driver, base_url: str) -> list:
    from test_suites.test_mobile_300_cases import run_300_appium_tests

    log.info("\n" + "="*60)
    log.info(f"  Running 300 Mobile Appium E2E Tests")
    log.info(f"  Target: {base_url}")
    log.info("="*60)

    raw_results = run_300_appium_tests(driver, base_url)
    results = []
    total = len(raw_results)

    for idx, r in enumerate(raw_results, 1):
        tc_id   = r["id"]
        tc_name = f"[{r['module']}] {r['name']} - {r['description']}"
        status  = r["status"]
        duration = r["duration"]
        error   = r["error"]

        if idx % 30 == 0 or idx == 1 or idx == total:
            log.info(f"[{idx}/{total}] {tc_id}: {r['name']} -> {status} ({duration}s)")

        results.append({
            "id":         tc_id,
            "name":       tc_name,
            "status":     status,
            "duration":   duration,
            "error":      error,
            "screenshot": "",
            "timestamp":  datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        })

    return results


# ─────────────────────────────────────────────────────────────────────────────
# Report Generation
# ─────────────────────────────────────────────────────────────────────────────
def generate_excel_report(results: list, driver_type: str, total_duration: float):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        passed  = sum(1 for r in results if r["status"] == "PASS")
        failed  = sum(1 for r in results if r["status"] == "FAIL")
        errored = sum(1 for r in results if r["status"] == "ERROR")
        total   = len(results)
        rate    = round((passed / total) * 100, 1) if total else 0

        # Styles
        font_hdr   = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        font_title = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
        font_bold  = Font(name="Calibri", size=11, bold=True)
        font_reg   = Font(name="Calibri", size=11)

        fill_navy  = PatternFill(start_color="1A3C6E", end_color="1A3C6E", fill_type="solid")
        fill_lb    = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        fill_pass  = PatternFill(start_color="375623", end_color="375623", fill_type="solid")
        fill_fail  = PatternFill(start_color="9C0006", end_color="9C0006", fill_type="solid")
        fill_err   = PatternFill(start_color="7B3F00", end_color="7B3F00", fill_type="solid")
        font_white = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

        thin = Border(
            left=Side(style='thin', color='BFBFBF'),
            right=Side(style='thin', color='BFBFBF'),
            top=Side(style='thin', color='BFBFBF'),
            bottom=Side(style='thin', color='BFBFBF')
        )

        def header_cell(ws, row, col, val):
            c = ws.cell(row=row, column=col, value=val)
            c.font = font_hdr; c.fill = fill_navy
            c.alignment = Alignment(horizontal="center", wrap_text=True)
            c.border = thin

        def data_cell(ws, row, col, val, wrap=False):
            c = ws.cell(row=row, column=col, value=val)
            c.font = font_reg; c.border = thin
            c.alignment = Alignment(wrap_text=wrap, vertical="top")
            return c

        wb = Workbook()

        # ── Sheet 1: Test Results ─────────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "Test Results"
        ws1.row_dimensions[1].height = 28
        headers = ["Test ID", "Test Name", "Status", "Duration (s)", "Timestamp", "Error Message", "Screenshot"]
        for col, h in enumerate(headers, 1):
            header_cell(ws1, 1, col, h)

        for ri, r in enumerate(results, 2):
            ws1.row_dimensions[ri].height = 40
            data_cell(ws1, ri, 1, r["id"])
            data_cell(ws1, ri, 2, r["name"])

            sc = ws1.cell(row=ri, column=3, value=r["status"])
            sc.alignment = Alignment(horizontal="center")
            sc.border = thin
            if r["status"] == "PASS":
                sc.fill = fill_pass; sc.font = font_white
            elif r["status"] == "FAIL":
                sc.fill = fill_fail; sc.font = font_white
            else:
                sc.fill = fill_err; sc.font = font_white

            data_cell(ws1, ri, 4, r["duration"])
            data_cell(ws1, ri, 5, r["timestamp"])
            data_cell(ws1, ri, 6, r["error"], wrap=True)
            data_cell(ws1, ri, 7, r["screenshot"])

        ws1.column_dimensions['A'].width = 10
        ws1.column_dimensions['B'].width = 38
        ws1.column_dimensions['C'].width = 10
        ws1.column_dimensions['D'].width = 14
        ws1.column_dimensions['E'].width = 20
        ws1.column_dimensions['F'].width = 42
        ws1.column_dimensions['G'].width = 40

        # ── Sheet 2: Test Summary ─────────────────────────────────────────────
        ws2 = wb.create_sheet("Test Summary")
        ws2.merge_cells("A1:C1")
        ws2["A1"] = "MOBILE E2E TEST SUMMARY"
        ws2["A1"].font = font_title
        ws2["A1"].fill = fill_navy
        ws2["A1"].alignment = Alignment(horizontal="center")

        summary_rows = [
            ("Execution Date",          datetime.now().strftime("%Y-%m-%d %H:%M")),
            ("GitHub Build Number",     GITHUB_RUN),
            ("Driver / Device",         driver_type),
            ("Target App URL",          BASE_URL),
            ("Total Test Cases",        total),
            ("Passed",                  passed),
            ("Failed",                  failed),
            ("Errors",                  errored),
            ("Pass Rate",               f"{rate}%"),
            ("Total Duration (s)",      round(total_duration, 2)),
        ]
        for ri, (label, val) in enumerate(summary_rows, 3):
            lc = ws2.cell(row=ri, column=1, value=label)
            lc.font = font_bold; lc.fill = fill_lb; lc.border = thin
            vc = ws2.cell(row=ri, column=2, value=val)
            vc.font = font_reg; vc.border = thin

        ws2.column_dimensions['A'].width = 28
        ws2.column_dimensions['B'].width = 35

        # ── Sheet 3: Execution Log ────────────────────────────────────────────
        ws3 = wb.create_sheet("Execution Log")
        log_headers = ["Timestamp", "Test ID", "Test Name", "Status", "Duration (s)", "Error"]
        for col, h in enumerate(log_headers, 1):
            header_cell(ws3, 1, col, h)
        for ri, r in enumerate(results, 2):
            data_cell(ws3, ri, 1, r["timestamp"])
            data_cell(ws3, ri, 2, r["id"])
            data_cell(ws3, ri, 3, r["name"])
            data_cell(ws3, ri, 4, r["status"])
            data_cell(ws3, ri, 5, r["duration"])
            data_cell(ws3, ri, 6, r["error"], wrap=True)
        ws3.column_dimensions['A'].width = 22
        ws3.column_dimensions['B'].width = 10
        ws3.column_dimensions['C'].width = 38
        ws3.column_dimensions['D'].width = 10
        ws3.column_dimensions['E'].width = 14
        ws3.column_dimensions['F'].width = 40

        # ── Sheet 4: Environment ──────────────────────────────────────────────
        ws4 = wb.create_sheet("Environment")
        env_data = [
            ("Framework",               "Appium 2.x / Selenium 4.x"),
            ("Driver Type",             driver_type),
            ("Target App URL",          BASE_URL),
            ("Appium Server URL",       APPIUM_URL),
            ("Device Name",             DEVICE_NAME),
            ("Android Version",         PLATFORM),
            ("Python Version",          f"{sys.version_info.major}.{sys.version_info.minor}.{sys.version_info.micro}"),
            ("Test Execution Mode",     "Mobile Chrome (Android)"),
            ("Report Generated",        datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ]
        ws4.merge_cells("A1:C1")
        ws4["A1"] = "TEST ENVIRONMENT DETAILS"
        ws4["A1"].font = font_title
        ws4["A1"].fill = fill_navy
        ws4["A1"].alignment = Alignment(horizontal="center")
        for ri, (label, val) in enumerate(env_data, 3):
            lc = ws4.cell(row=ri, column=1, value=label)
            lc.font = font_bold; lc.fill = fill_lb; lc.border = thin
            vc = ws4.cell(row=ri, column=2, value=val)
            vc.font = font_reg; vc.border = thin
        ws4.column_dimensions['A'].width = 28
        ws4.column_dimensions['B'].width = 40

        path = os.path.join(EXCEL_DIR, "Appium_Mobile_Test_Report.xlsx")
        wb.save(path)
        log.info(f"[Excel] Report saved: {path}")
        return path

    except ImportError:
        log.warning("openpyxl not installed — skipping Excel report")
        return ""
    except Exception as e:
        log.error(f"Excel report generation failed: {e}")
        return ""


def generate_html_report(results: list, driver_type: str, total_duration: float) -> str:
    passed  = sum(1 for r in results if r["status"] == "PASS")
    failed  = sum(1 for r in results if r["status"] == "FAIL")
    errored = sum(1 for r in results if r["status"] == "ERROR")
    total   = len(results)
    rate    = round((passed / total) * 100, 1) if total else 0
    overall_color = "#22c55e" if rate == 100 else "#ef4444" if rate < 50 else "#f59e0b"

    rows = ""
    for r in results:
        badge_cls = {"PASS": "badge-pass", "FAIL": "badge-fail"}.get(r["status"], "badge-error")
        err = r["error"].replace("<", "&lt;").replace(">", "&gt;") if r["error"] else "—"
        rows += f"""
        <tr>
          <td>{r['id']}</td>
          <td>{r['name']}</td>
          <td><span class="badge {badge_cls}">{r['status']}</span></td>
          <td>{r['duration']}s</td>
          <td class="error-cell">{err}</td>
          <td>{r['timestamp']}</td>
        </tr>"""

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Appium Mobile E2E — Execution Report</title>
<style>
  :root {{
    --bg: #0f172a; --card: #1e293b; --border: #334155;
    --text: #f1f5f9; --muted: #94a3b8;
    --green: #22c55e; --red: #ef4444; --amber: #f59e0b; --blue: #3b82f6;
  }}
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ background: var(--bg); color: var(--text); font-family: 'Segoe UI', system-ui, sans-serif; padding: 2rem; }}
  h1 {{ font-size: 1.8rem; font-weight: 700; margin-bottom: 0.25rem; }}
  .meta {{ color: var(--muted); font-size: 0.875rem; margin-bottom: 2rem; }}
  .grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(160px,1fr)); gap: 1rem; margin-bottom: 2rem; }}
  .card {{ background: var(--card); border: 1px solid var(--border); border-radius: 12px; padding: 1.25rem; text-align: center; }}
  .card .val {{ font-size: 2.5rem; font-weight: 800; }}
  .card .lbl {{ font-size: 0.8rem; color: var(--muted); margin-top: 0.25rem; text-transform: uppercase; letter-spacing: .05em; }}
  .score-bar {{ height: 8px; border-radius: 4px; background: #1e293b; margin-bottom: 2rem; }}
  .score-fill {{ height: 100%; border-radius: 4px; background: {overall_color}; width: {rate}%; transition: width .5s; }}
  table {{ width: 100%; border-collapse: collapse; background: var(--card); border-radius: 12px; overflow: hidden; }}
  thead {{ background: #0f172a; }}
  th {{ padding: .75rem 1rem; text-align: left; font-size: .8rem; text-transform: uppercase; letter-spacing:.05em; color: var(--muted); }}
  td {{ padding: .7rem 1rem; border-top: 1px solid var(--border); font-size: .875rem; vertical-align: top; }}
  .badge {{ display: inline-block; padding: .2rem .6rem; border-radius: 6px; font-size: .75rem; font-weight: 700; }}
  .badge-pass  {{ background: #14532d; color: var(--green); }}
  .badge-fail  {{ background: #450a0a; color: var(--red); }}
  .badge-error {{ background: #451a03; color: var(--amber); }}
  .error-cell {{ color: var(--red); font-family: monospace; font-size: .8rem; max-width: 360px; word-break: break-word; }}
  .device-bar {{ background: var(--card); border: 1px solid var(--border); border-radius: 10px; padding: 1rem 1.5rem; margin-bottom: 2rem; font-size: .875rem; color: var(--muted); }}
  .device-bar span {{ color: var(--text); font-weight: 600; }}
</style>
</head>
<body>
<h1>Appium Mobile E2E — Execution Report</h1>
<p class="meta">Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} &nbsp;|&nbsp; Build: {GITHUB_RUN}</p>

<div class="device-bar">
  Driver: <span>{driver_type}</span> &nbsp;|&nbsp;
  Target: <span>{BASE_URL}</span> &nbsp;|&nbsp;
  Device: <span>{DEVICE_NAME} (Android {PLATFORM})</span> &nbsp;|&nbsp;
  Duration: <span>{round(total_duration,1)}s</span>
</div>

<div class="grid">
  <div class="card"><div class="val">{total}</div><div class="lbl">Total Tests</div></div>
  <div class="card"><div class="val" style="color:var(--green)">{passed}</div><div class="lbl">Passed</div></div>
  <div class="card"><div class="val" style="color:var(--red)">{failed + errored}</div><div class="lbl">Failed</div></div>
  <div class="card"><div class="val" style="color:{overall_color}">{rate}%</div><div class="lbl">Pass Rate</div></div>
</div>

<div class="score-bar"><div class="score-fill"></div></div>

<table>
<thead>
  <tr>
    <th>ID</th><th>Test Name</th><th>Status</th><th>Duration</th><th>Error</th><th>Timestamp</th>
  </tr>
</thead>
<tbody>{rows}</tbody>
</table>
</body>
</html>"""

    path = os.path.join(HTML_DIR, "mobile_execution_report.html")
    with open(path, "w", encoding="utf-8") as f:
        f.write(html)
    log.info(f"[HTML] Report saved: {path}")
    return path


def generate_markdown_summary(results: list, driver_type: str, total_duration: float) -> str:
    passed  = sum(1 for r in results if r["status"] == "PASS")
    failed  = sum(1 for r in results if r["status"] == "FAIL")
    errored = sum(1 for r in results if r["status"] == "ERROR")
    total   = len(results)
    rate    = round((passed / total) * 100, 1) if total else 0

    rows = "\n".join(
        f"| {r['id']} | {r['name']} | {'OK' if r['status']=='PASS' else 'FAIL' if r['status']=='FAIL' else 'ERR'} {r['status']} | {r['duration']}s |"
        for r in results
    )

    md = f"""# Appium Mobile E2E Test Report

- **Date**: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
- **Build**: {GITHUB_RUN}
- **Driver**: {driver_type}
- **Target**: {BASE_URL}

## Summary

| Metric | Value |
|---|---|
| Total Tests | {total} |
| Passed | {passed} |
| Failed | {failed} |
| Errors | {errored} |
| Pass Rate | **{rate}%** |
| Total Duration | {round(total_duration, 1)}s |

## Test Results

| ID | Test Name | Status | Duration |
|---|---|---|---|
{rows}
"""

    path = os.path.join(SUMMARY_DIR, "mobile_summary.md")
    with open(path, "w", encoding="utf-8") as f:
        f.write(md)
    log.info(f"[MD] Summary saved: {path}")
    return path


# ─────────────────────────────────────────────────────────────────────────────
# Main Entrypoint
# ─────────────────────────────────────────────────────────────────────────────
def main():
    log.info("=" * 60)
    log.info("  VoiceMail AI — Appium Mobile E2E Test Runner")
    log.info("=" * 60)
    log.info(f"  Target URL   : {BASE_URL}")
    log.info(f"  Appium URL   : {APPIUM_URL}")
    log.info(f"  Device       : {DEVICE_NAME} (Android {PLATFORM})")
    log.info(f"  Build Number : {GITHUB_RUN}")

    if not wait_for_server(BASE_URL):
        log.error("Aborting — target server not reachable.")
        sys.exit(1)

    driver, driver_type = create_appium_driver()
    log.info(f"  Driver Mode  : {driver_type}")

    suite_start = time.time()
    results = []
    try:
        results = run_all_tests(driver, BASE_URL)
    finally:
        try:
            driver.quit()
        except Exception:
            pass

    total_duration = time.time() - suite_start

    passed  = sum(1 for r in results if r["status"] == "PASS")
    failed  = sum(1 for r in results if r["status"] in ("FAIL", "ERROR"))
    total   = len(results)
    rate    = round((passed / total) * 100, 1) if total else 0

    log.info("\n" + "="*60)
    log.info(f"  RESULTS: {passed}/{total} passed ({rate}%)  |  Duration: {round(total_duration,1)}s")
    log.info("="*60 + "\n")

    # Generate all reports
    generate_excel_report(results, driver_type, total_duration)
    generate_html_report(results, driver_type, total_duration)
    generate_markdown_summary(results, driver_type, total_duration)

    log.info(f"[DONE] All reports saved to: {os.path.abspath(RESULTS_ROOT)}/")
    log.info(f"   Excel  -> {EXCEL_DIR}/Appium_Mobile_Test_Report.xlsx")
    log.info(f"   HTML   -> {HTML_DIR}/mobile_execution_report.html")
    log.info(f"   MD     -> {SUMMARY_DIR}/mobile_summary.md")
    log.info(f"   Shots  -> {SCREENSHOT_DIR}/")
    log.info(f"   Logs   -> {LOGS_DIR}/appium_e2e.log")

    sys.exit(0 if failed == 0 else 1)


if __name__ == "__main__":
    main()
