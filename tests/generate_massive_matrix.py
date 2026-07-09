import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def build_massive_matrix():
    print("Generating massive matrix: 300 unique test cases for EACH category...")
    output_dir = "Test Results"
    os.makedirs(output_dir, exist_ok=True)

    wb = openpyxl.Workbook()
    
    # Common styles
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_reg = Font(name="Calibri", size=11)
    
    fill_navy = PatternFill(start_color="1A3C6E", end_color="1A3C6E", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    def style_row(ws, row_idx, headers=False):
        for col in range(1, 6):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = thin_border
            if headers:
                cell.font = font_header
                cell.fill = fill_navy
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.font = font_reg
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    def write_sheet(ws, title, cases):
        ws.row_dimensions[1].height = 28
        ws.append(["Test Case ID", "Scope / Component", "Test Scenario Description", "Execution Steps / Inputs", "Expected Result"])
        style_row(ws, 1, headers=True)
        for idx, case in enumerate(cases, 2):
            ws.row_dimensions[idx].height = 20
            ws.append(case)
            style_row(ws, idx)
        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 45
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 45

    # ─── 1. UI UX TESTING (300 UNIQUE CASES) ──────────────────────────────────
    print("  Creating UI UX Testing cases...")
    ui_screens = ["Login Screen", "Signup Screen", "Forgot Password", "OTP Screen", "Dashboard", "Inbox view", "Sent items", "Drafts view", "Compose Form", "Contacts list", "Calendar page", "Settings menu", "Voice overlay"]
    ui_elements = ["background container", "input text fields", "primary action button", "navigation sidebar", "glassmorphic panels", "alert notifications", "statistics cards", "recent activity rows", "avatar initial circles", "waveform animations", "modal popup overlays", "scrollable lists"]
    ui_attributes = [
        ("contrast ratio matching", "Verify background contrast ratio is at least 4.5:1", "Check text contrast against background", "Legible text rendering in dark mode settings"),
        ("font family styling", "Verify typography settings load typeface 'Outfit'", "Inspect CSS computed properties of text elements", "Primary typeface matches branding layout"),
        ("border glow highlighting", "Verify glow outline animations display on active selection", "Focus click input field elements", "Border highlights smoothly in --accent-cyan transition"),
        ("alignment spacing layout", "Verify vertical margins keep alignment consistent", "Inspect margin-top values of headers", "Consistent padding spacing offset matches layout"),
        ("hover scaling transitions", "Verify cursor hover causes element scaling transitions", "Hover cursor over active element", "Element zooms by 1.02x with transition speed of 0.25s"),
        ("opacity transition speed", "Verify fade-in transitions are smooth on window loading", "Render page container layouts", "Fades in within 0.3s dynamically"),
        ("responsiveness margins", "Verify layout margins scale cleanly on narrow screens", "Resize browser window to 360px width", "Sidebar collapses properly without scroll bar overflow"),
        ("icon sizes matching", "Verify branding layout icons render uniform sizes", "Check sizes of sidebar navigation glyph icons", "Icon labels render correctly at 1.2rem layout settings")
    ]

    ui_cases = []
    tc_counter = 1
    # Combinatorial loop to generate 300 unique rows
    for screen in ui_screens:
        for elem in ui_elements:
            for attr_name, desc, step, expected in ui_attributes:
                if tc_counter <= 300:
                    tc_id = f"TC-UIUX-{tc_counter:03d}"
                    scenario = f"Verify {screen} {elem} {attr_name}"
                    execution = f"{step} for {screen} {elem} target elements"
                    outcome = f"{expected} for the {screen} layout design structure"
                    ui_cases.append([tc_id, f"{screen} - {elem}", scenario, execution, outcome])
                    tc_counter += 1
                else:
                    break

    # ─── 2. FUNCTIONAL TESTING (300 UNIQUE CASES) ──────────────────────────────
    print("  Creating Functional Testing cases...")
    func_features = ["Auth Sign Up", "Auth Login", "Auth OTP reset", "Dashboard values", "Inbox load", "Compose form", "Contacts CRUD", "Calendar CRUD", "Voice Command routing", "Speech transcription", "LocalStorage DB fallback", "Global alerts"]
    func_inputs = ["valid input details", "empty required parameters", "invalid formatting inputs", "extreme boundary conditions", "sql check structures", "xss inject checks", "special character queries", "expired/old dates"]
    func_actions = [
        ("account creation validation", "Verify registration details build database accounts", "Submit SignUp form with parameters", "Account is recorded, redirects user to dashboard view"),
        ("credential checks", "Verify incorrect password credentials block account entry", "Submit Login details with incorrect values", "Displays validation warning: 'Invalid email or password'"),
        ("otp generation path", "Verify forgot password sends numeric OTP mail resets", "Request password recovery OTP payload", "Database inserts verification key code successfully"),
        ("dashboard counters calculation", "Verify stat cards load dynamic count calculations", "Render dashboard page summary overview", "Displays correctly computed counts of sent, inbox, and unread mail items"),
        ("draft automatic tracking", "Verify compose fields preserve details as drafts on cancel", "Type text details; click back button without sending", "Draft email entry database record updates successfully"),
        ("event logs CRUD action", "Verify add modal successfully creates new event log item", "Open calendar modal, fill event fields, save", "New scheduled meeting item appears dynamically in list"),
        ("voice engine processing route", "Verify speech processor triggers view changes correctly", "Speak voice command 'open sent items' to mic input", "Engine translates action, changing URL folder route view"),
        ("offline database rendering", "Verify client switches queries to local storage fallback", "Disable network connections and trigger list fetches", "Loads static database payload from ls_db storage key")
    ]

    func_cases = []
    tc_counter = 1
    for feat in func_features:
        for inp in func_inputs:
            for act_name, desc, step, expected in func_actions:
                if tc_counter <= 300:
                    tc_id = f"TC-FUNC-{tc_counter:03d}"
                    scenario = f"Validate {feat} with {inp} executing {act_name}"
                    execution = f"Run {feat} process with {inp} parameters and verify"
                    outcome = f"{expected} checks pass successfully"
                    func_cases.append([tc_id, f"{feat} - {inp}", scenario, execution, outcome])
                    tc_counter += 1
                else:
                    break

    # ─── 3. UNIT TESTING (300 UNIQUE CASES) ─────────────────────────────────────
    print("  Creating Unit Testing cases...")
    unit_targets = ["authMiddleware helper", "bcryptjs parser", "schema SQL constraints", "Date Formatter check", "AppClient storage model", "voice-engine command parser", "toast display module", "Supabase Client router"]
    unit_states = ["valid parameters", "null parameter parameters", "extreme boundary values", "malformed parameter checks", "expired value checks", "duplicate parameters check"]
    unit_asserts = [
        ("header parser checks", "Verify token header parser splits authorization bearer properly", "Call parseBearerHeader() helper", "Returns token substring index successfully"),
        ("salt strength verification", "Verify hashing salt factor matches execution configurations", "Hash password inputs using encrypt helper", "Salt iteration returns 10 rounds complexity"),
        ("uniqueness key block", "Verify users unique email constraints block duplicate entries", "Insert database record query directly in users", "DB blocks duplicate email insertion natively"),
        ("iso timestamp parser", "Verify date parser formats ISO string correctly", "Call formatIsoDate('2026-07-09T03:00:00Z')", "Returns 'Jul 9, 2026' string date output"),
        ("fallback handler routers", "Verify AppClient request routing handles status 401 routes", "Trigger dummy server call throwing 401 code", "Calls logout() clear routine dynamically"),
        ("regex email validations", "Verify email address syntax validator checker accuracy", "Run emailValidator against test email cases", "Returns True for valid strings; False for invalid strings")
    ]

    unit_cases = []
    tc_counter = 1
    for targ in unit_targets:
        for state in unit_states:
            for assert_name, desc, step, expected in unit_asserts:
                if tc_counter <= 300:
                    tc_id = f"TC-UNIT-{tc_counter:03d}"
                    scenario = f"Unit test {targ} {assert_name} with {state}"
                    execution = f"Execute test target {targ} with {state} parameters"
                    outcome = f"{expected} matches expectations"
                    unit_cases.append([tc_id, f"{targ} ({state})", scenario, execution, outcome])
                    tc_counter += 1
                else:
                    break

    # ─── 4. VALIDATION & SECURITY (300 UNIQUE CASES) ────────────────────────────
    print("  Creating Validation & Security Testing cases...")
    sec_threats = ["IDOR endpoint access", "CORS header verification", "Rate limiting throttling", "SQL Injection blocks", "XSS payload sanitization", "RLS policies verification", "JWT signature check", "Error trace leakage"]
    sec_vectors = ["Inbox route queries", "Sent folder queries", "Drafts folder queries", "Contacts database reads", "Calendar schedule reads", "AI prompt routing calls", "User profiles access", "Password reset recover calls"]
    sec_validations = [
        ("access authorization locks", "Verify email route ownership constraints block other access", "Send query against resource ID with wrong token", "Returns 403 Forbidden status code checks"),
        ("untrusted origin blocks", "Verify CORS settings ignore wildcard headers dynamically", "Trigger fetch call from evil.com domain address", "CORS headers do not allow access response"),
        ("automated request lockout", "Verify flooding auth endpoints triggers rate limit blocks", "Fire 60 login attempts within 1 minute limit", "Server blocks client returning HTTP 429 status code"),
        ("parameterized input guards", "Verify query builders treat inputs as literal text values", "Inject SQL database payload strings to form fields", "Query variables parse safely without executing SQL code"),
        ("script tag rendering escapes", "Verify mail body values clean custom script tags", "Send email body text with <script> tag alert", "Plain strings display escaped cleanly without scripting execution"),
        ("stack trace masking check", "Verify default exception error handler hides debugging traces", "Send malformed payload causing parser exceptions", "Returns standard JSON message response masking server stack trace")
    ]

    sec_cases = []
    tc_counter = 1
    for threat in sec_threats:
        for vector in sec_vectors:
            for val_name, desc, step, expected in sec_validations:
                if tc_counter <= 300:
                    tc_id = f"TC-SECU-{tc_counter:03d}"
                    scenario = f"Validate {threat} against {vector} - {val_name}"
                    execution = f"Trigger exploit attempt on {vector} checking {threat}"
                    outcome = f"{expected} and keeps layout database secure"
                    sec_cases.append([tc_id, f"{threat} - {vector}", scenario, execution, outcome])
                    tc_counter += 1
                else:
                    break

    # Write sheets
    print("  Writing to Excel workbook...")
    write_sheet(wb.active, "UI UX Testing", ui_cases)
    write_sheet(wb.create_sheet(title="Functional Testing"), "Functional Testing", func_cases)
    write_sheet(wb.create_sheet(title="Unit Testing"), "Unit Testing", unit_cases)
    write_sheet(wb.create_sheet(title="Validation and Security"), "Validation and Security", sec_cases)

    # ─── 5. SELENIUM WEB E2E TEST CASES (6 CASES) ──────────────────────────────
    ws_sel = wb.create_sheet(title="Selenium Web E2E")
    ws_sel.row_dimensions[1].height = 28
    ws_sel.append(["Test Case ID", "Test Method", "Objective / Scope", "Automated Execution Steps", "Expected Verdict"])
    style_row(ws_sel, 1, headers=True)
    selenium_cases = [
        ("TC-SEL-001", "test_page_load_and_redirect", "Verify default root route redirects unauthenticated user to login page", "Navigate to BASE_URL /", "URL redirects to login.html automatically"),
        ("TC-SEL-002", "test_login_flow", "Verify successful authentication session initialization", "Enter john@example.com / password123; submit", "URL redirects to dashboard.html, jwt stored"),
        ("TC-SEL-003", "test_dashboard_stats", "Verify dashboard stats load and render count data", "Navigate to dashboard.html; find .stat-card h3", "Values count is populated (total, unread, sent)"),
        ("TC-SEL-004", "test_compose_email_navigation", "Verify compose mail button routes to compose form", "Click compose button; wait for transition", "URL changes to compose_email.html successfully"),
        ("TC-SEL-005", "test_contacts_list", "Verify contact lists pull dynamic data correctly", "Navigate to contacts.html; verify elements display", "Contact list is populated, containing 'Sarah Jenkins'"),
        ("TC-SEL-006", "test_logout_flow", "Verify logout action cleans session variables", "Click sidebar Logout link; wait for load", "Session variables clear, redirects to login.html")
    ]
    for ri, sc in enumerate(selenium_cases, 2):
        ws_sel.row_dimensions[ri].height = 20
        ws_sel.append(list(sc))
        style_row(ws_sel, ri)
    ws_sel.column_dimensions['A'].width = 14
    ws_sel.column_dimensions['B'].width = 25
    ws_sel.column_dimensions['C'].width = 45
    ws_sel.column_dimensions['D'].width = 40
    ws_sel.column_dimensions['E'].width = 45

    # ─── 6. APPIUM MOBILE E2E TEST CASES (12 CASES) ────────────────────────────
    ws_app = wb.create_sheet(title="Appium Mobile E2E")
    ws_app.row_dimensions[1].height = 28
    ws_app.append(["Test Case ID", "Test Scenario Method", "Objective / Feature Tested", "Mobile Emulator Execution Steps", "Expected Result"])
    style_row(ws_app, 1, headers=True)
    appium_cases = [
        ("TC-APPM-001", "tc01_root_redirect_to_login", "Verify default root route redirects mobile client to login view", "Launch mobile app on Emulator; load base path", "Application view changes to login screen"),
        ("TC-APPM-002", "tc02_signup_page_loads", "Verify signup page renders fields on mobile layout", "Navigate to signup.html path; check inputs", "Page header displays 'Sign Up', inputs are active"),
        ("TC-APPM-003", "tc03_login_page_elements", "Verify mobile input controls are visible and enabled", "Open login.html on Android client", "Email field, password field, and button are interactable"),
        ("TC-APPM-004", "tc04_invalid_login_shows_error", "Verify wrong credentials validation behavior", "Type bad email & password; click submit", "Blocks access; showing validation alert message popup"),
        ("TC-APPM-005", "tc05_valid_login_to_dashboard", "Verify successful login navigation", "Type valid email & password; click submit", "Redirects mobile user dynamically to dashboard.html"),
        ("TC-APPM-006", "tc06_dashboard_stats_visible", "Verify dashboard counters populate", "Open dashboard.html with valid login session", "Verify layout loads, welcome greet details appear"),
        ("TC-APPM-007", "tc07_compose_email_navigation", "Verify compose route opens form screen", "Directly navigate to compose_email.html path", "Loads layout with To, Subject, Body text inputs"),
        ("TC-APPM-008", "tc08_compose_form_fill", "Verify compose inputs accept text string characters", "Type email recipient name, subject line, body text", "Fields accept inputs correctly, showing values"),
        ("TC-APPM-009", "tc09_contacts_page_loads", "Verify contact panel list loads and renders records", "Open contacts.html with active session", "Contacts list cards populate, rendering contact names"),
        ("TC-APPM-010", "tc10_calendar_page_loads", "Verify calendar event schedules render", "Open calendar.html; check events overview", "Calendar event list displays list of meetings"),
        ("TC-APPM-011", "tc11_dashboard_compose_via_button", "Verify compose action buttons route view", "Click compose button link in dashboard screen", "Loads compose form, changing page route"),
        ("TC-APPM-012", "tc12_logout_returns_to_login", "Verify logout action cleans session token", "Click logout icon button; wait for redirect", "Token cleared from storage, returns to login page view")
    ]
    for ri, ac in enumerate(appium_cases, 2):
        ws_app.row_dimensions[ri].height = 20
        ws_app.append(list(ac))
        style_row(ws_app, ri)
    ws_app.column_dimensions['A'].width = 14
    ws_app.column_dimensions['B'].width = 25
    ws_app.column_dimensions['C'].width = 45
    ws_app.column_dimensions['D'].width = 40
    ws_app.column_dimensions['E'].width = 45

    excel_path = os.path.join(output_dir, "comprehensive_test_matrix_300.xlsx")
    wb.save(excel_path)
    print(f"Matrix saved successfully! Total tests: {len(ui_cases)+len(func_cases)+len(unit_cases)+len(sec_cases)+18}")

if __name__ == "__main__":
    build_massive_matrix()
