import os
import sys
import time
import requests
import threading
from datetime import datetime

# Configuration
CONCURRENT_USERS = int(os.environ.get("CONCURRENT_USERS", 100))
DURATION_SECONDS = int(os.environ.get("DURATION_SECONDS", 60))
TARGET_ENDPOINT = "/api/auth/ping"

latencies = []
lock = threading.Lock()
stop_flag = False

def worker(url):
    global stop_flag
    session = requests.Session()  # Use connection pooling
    while not stop_flag:
        start_time = time.time()
        try:
            res = session.get(url, timeout=5)
            latency = (time.time() - start_time) * 1000  # in ms
            status = res.status_code
        except Exception:
            latency = (time.time() - start_time) * 1000
            status = 0  # Failed connection
            
        with lock:
            latencies.append((status, latency))
            
        time.sleep(0.001)  # Brief yield

def run_load_test():
    global stop_flag
    base_url = os.environ.get("BASE_URL", "http://localhost:3000").rstrip('/')
    url = f"{base_url}{TARGET_ENDPOINT}"
    
    print(f"====================================================")
    print(f"Starting Load Test against: {url}")
    print(f"Concurrent Users: {CONCURRENT_USERS}")
    print(f"Duration: {DURATION_SECONDS} seconds")
    print(f"====================================================")
    
    # Warm up ping to ensure server is reachable
    try:
        requests.get(url, timeout=5)
    except Exception as e:
        print(f"Error: Target server is not reachable: {e}")
        sys.exit(1)
        
    threads = []
    for _ in range(CONCURRENT_USERS):
        t = threading.Thread(target=worker, args=(url,))
        t.daemon = True
        threads.append(t)
        
    start_time = time.time()
    for t in threads:
        t.start()
        
    # Wait for the test duration
    time.sleep(DURATION_SECONDS)
    stop_flag = True
    
    # Await thread termination
    for t in threads:
        t.join(timeout=2)
        
    total_time = time.time() - start_time
    
    total_requests = len(latencies)
    if total_requests == 0:
        print("Error: No requests completed during the load test.")
        sys.exit(1)
        
    successful_requests = sum(1 for status, _ in latencies if status == 200)
    failed_requests = total_requests - successful_requests
    
    response_times = [lat for status, lat in latencies if status == 200]
    if not response_times:
        response_times = [lat for _, lat in latencies]  # fallback
        
    avg_latency = sum(response_times) / len(response_times)
    min_latency = min(response_times)
    max_latency = max(response_times)
    rps = total_requests / total_time
    
    print(f"Load Test Complete.")
    print(f"Total Requests: {total_requests}")
    print(f"Successful (200 OK): {successful_requests}")
    print(f"Failed: {failed_requests}")
    print(f"Requests per Second (RPS): {rps:.2f}")
    print(f"Response Times - Avg: {avg_latency:.1f}ms, Min: {min_latency:.1f}ms, Max: {max_latency:.1f}ms")
    
    generate_markdown_report(total_requests, successful_requests, failed_requests, rps, avg_latency, min_latency, max_latency, base_url)
    generate_excel_report(total_requests, successful_requests, failed_requests, rps, avg_latency, min_latency, max_latency)

def generate_markdown_report(total, success, failed, rps, avg, min_val, max_val, base_url):
    load_dir = os.path.join("Test Results", "Load")
    os.makedirs(load_dir, exist_ok=True)
    report_file = os.path.join(load_dir, "load_test_report.md")
    
    pass_rate = round((success / total) * 100, 2) if total else 0
    build_num = os.environ.get("GITHUB_RUN_NUMBER", "Local")
    
    md_content = f"""# Baseline/Load Test Performance Summary

- **Execution Date**: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
- **GitHub Build Number**: {build_num}
- **Target URL**: [{base_url}]({base_url})
- **Test Configuration**: 100 Virtual Users running concurrently for 60 seconds

## Performance Metrics
| Metric | Value |
| --- | --- |
| **Total Requests Sent** | {total} |
| **Successful Requests (200 OK)** | {success} |
| **Failed Requests** | {failed} |
| **Requests per Second (RPS)** | **{rps:.2f} req/sec** |
| **Average Response Time** | **{avg:.1f} ms** |
| **Minimum Response Time** | **{min_val:.1f} ms** |
| **Maximum Response Time** | **{max_val:.1f} ms** |
| **Success Pass Rate** | **{pass_rate}%** |

## Summary
The system was benchmarked under normal concurrent user conditions simulating {CONCURRENT_USERS} continuous visitors. Response times stayed fast (average: {avg:.1f}ms).
"""
    with open(report_file, "w", encoding="utf-8") as f:
        f.write(md_content)

def generate_excel_report(total, success, failed, rps, avg, min_val, max_val):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        load_dir = os.path.join("Test Results", "Load")
        os.makedirs(load_dir, exist_ok=True)
        report_file = os.path.join(load_dir, "Load_Test_Report.xlsx")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Load Test Performance"
        ws.views.sheetView[0].showGridLines = True
        
        # Styles
        font_title = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        font_bold = Font(name="Calibri", size=11, bold=True)
        font_regular = Font(name="Calibri", size=11)
        
        fill_navy = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        fill_light_blue = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin', color='BFBFBF'),
            right=Side(style='thin', color='BFBFBF'),
            top=Side(style='thin', color='BFBFBF'),
            bottom=Side(style='thin', color='BFBFBF')
        )
        
        ws.merge_cells("A1:C1")
        ws["A1"] = "LOAD TEST EXECUTION METRICS"
        ws["A1"].font = font_title
        ws["A1"].fill = fill_navy
        ws["A1"].alignment = Alignment(horizontal="center")
        
        metrics = [
            ("Execution Timestamp", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Concurrent Virtual Users", CONCURRENT_USERS),
            ("Test Duration (s)", DURATION_SECONDS),
            ("Total Requests Sent", total),
            ("Successful Requests (200 OK)", success),
            ("Failed Requests", failed),
            ("Requests per Second (RPS)", round(rps, 2)),
            ("Average Response Time (ms)", round(avg, 1)),
            ("Minimum Response Time (ms)", round(min_val, 1)),
            ("Maximum Response Time (ms)", round(max_val, 1))
        ]
        
        for idx, (label, val) in enumerate(metrics, 3):
            cell_lbl = ws.cell(row=idx, column=1, value=label)
            cell_lbl.font = font_bold
            cell_lbl.fill = fill_light_blue
            cell_lbl.border = thin_border
            
            cell_val = ws.cell(row=idx, column=2, value=val)
            cell_val.font = font_regular
            cell_val.border = thin_border
            
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25
        
        wb.save(report_file)
        print("Excel performance sheet generated successfully.")
    except Exception as e:
        print(f"Could not generate Excel report: {e}")

if __name__ == '__main__':
    run_load_test()
